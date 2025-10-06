import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from thefuzz import fuzz, process

st.set_page_config(page_title="Excel Analyzer Pro", page_icon="📊", layout="wide")
st.title("📊 Excel Analyzer Pro - Analyse intelligente de contrats")
st.markdown("### Embellissez, analysez et recherchez dans vos fichiers Excel")

# ==================== FONCTIONS ====================

def clean_data(df):
    df = df.dropna(how='all').dropna(axis=1, how='all')
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].astype(str).str.strip()
    return df.replace('nan', '').fillna('')

def parse_nl_query(query, df):
    filters = {}
    q = query.lower()
    if any(w in q for w in ['ko','échec','erreur','rejet']): filters['statut'] = 'KO'
    elif any(w in q for w in ['ok','réussi','succès']): filters['statut'] = 'OK'
    if 'Code_Unite' in df.columns:
        for ag in df['Code_Unite'].unique():
            if str(ag).lower() in q: filters['agence'] = ag; break
    mois = {'janvier':1,'jan':1,'février':2,'fev':2,'mars':3,'avril':4,'mai':5,'juin':6,'juillet':7,'août':8,'aout':8,'septembre':9,'sept':9,'octobre':10,'novembre':11,'décembre':12,'dec':12}
    for n,m in mois.items():
        if n in q: filters['mois'] = m; break
    if 'initial' in q: filters['init_avenant'] = 'Initial'
    elif 'avenant' in q: filters['init_avenant'] = 'Avenant'
    return filters

def fuzzy_search(query, df, col, lim=10):
    if col not in df.columns: return []
    vals = [v for v in df[col].dropna().astype(str).unique() if v.strip()]
    if not vals or not query.strip(): return []
    return [(m[0],m[1]) for m in process.extract(query,vals,limit=lim,scorer=fuzz.token_sort_ratio) if m[1]>50]

def calc_score(row, query, filters):
    score = 0
    if 'Contrat' in row.index:
        if query.lower() in str(row['Contrat']).lower(): score += 100
        else: score += fuzz.partial_ratio(query.lower(), str(row['Contrat']).lower()) * 0.5
    if filters.get('agence') and row.get('Code_Unite') == filters['agence']: score += 50
    if filters.get('statut'):
        if filters['statut']=='KO' and str(row.get('Statut_Final','')).upper()!='OK': score += 50
        elif filters['statut']=='OK' and str(row.get('Statut_Final','')).upper()=='OK': score += 50
    if filters.get('mois'):
        try:
            if pd.to_datetime(row.get('Date_Integration')).month == filters['mois']: score += 40
        except: pass
    return score

def get_suggestions(inp, df, lim=5):
    if not inp or len(inp)<2: return []
    sugg = []
    if 'Contrat' in df.columns:
        for c in df['Contrat'].dropna().astype(str)[df['Contrat'].astype(str).str.contains(inp,case=False,na=False)].head(lim):
            sugg.append({'type':'📄 Contrat','value':c,'score':fuzz.partial_ratio(inp.lower(),c.lower())})
    if 'Code_Unite' in df.columns:
        for ag in df['Code_Unite'].unique():
            if inp.lower() in str(ag).lower(): sugg.append({'type':'🏢 Agence','value':ag,'score':100})
    if 'ko' in inp.lower(): sugg.append({'type':'❌ Statut','value':'KO','score':100})
    if 'ok' in inp.lower(): sugg.append({'type':'✅ Statut','value':'OK','score':100})
    return sorted(sugg, key=lambda x:x['score'], reverse=True)[:lim]

def style_ws(ws):
    hf = PatternFill(start_color="366092",end_color="366092",fill_type="solid")
    for c in ws[1]:
        c.fill = hf
        c.font = Font(bold=True,color="FFFFFF",size=11)
        c.alignment = Alignment(horizontal='center',vertical='center')
        c.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    for i,row in enumerate(ws.iter_rows(min_row=2,max_row=ws.max_row),2):
        fill = PatternFill(start_color="F2F2F2" if i%2==0 else "FFFFFF",fill_type="solid")
        for c in row:
            c.fill = fill
            c.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or '')) for c in col)+2, 50)
    ws.freeze_panes = 'A2'

def create_excel(df):
    """Crée Excel ULTRA-DÉTAILLÉ avec 7 onglets complets"""
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    
    # ONGLET 1: Données nettoyées
    ws1 = wb.create_sheet('Données nettoyées')
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    style_ws(ws1)
    
    # ONGLET 2: Vue d'ensemble
    total = len(df)
    ok = len(df[df['Statut_Final'].str.upper()=='OK'])
    ko = total - ok
    taux = round(ok/total*100,2) if total else 0
    init = len(df[df['Initial/Avenant'].str.contains('Initial',case=False,na=False)])
    aven = len(df[df['Initial/Avenant'].str.contains('Avenant',case=False,na=False)])
    
    ws2 = wb.create_sheet('Vue d ensemble')
    vue = [['Métrique','Valeur'],
           ['Nombre total de contrats',total],
           ['Nombre de contrats OK',ok],
           ['Nombre de contrats KO',ko],
           ['Taux de réussite (%)',f'{taux}%'],
           ['Nombre de contrats initiaux',init],
           ['Nombre d\'avenants',aven],
           ['Nombre d\'agences',df['Code_Unite'].nunique() if 'Code_Unite' in df.columns else 0]]
    for r in vue: ws2.append(r)
    style_ws(ws2)
    
    # ONGLET 3: Analyse par agence ULTRA-DÉTAILLÉE
    if 'Code_Unite' in df.columns and 'Statut_Final' in df.columns:
        ws3 = wb.create_sheet('Analyse par agence')
        r = 1
        
        # Titre
        ws3.cell(r,1,'ANALYSE COMPLÈTE PAR AGENCE (CODE_UNITE)').font = Font(bold=True,size=14,color="366092")
        ws3.merge_cells(f'A{r}:G{r}')
        r += 2
        
        # Calculer métriques
        ag_metrics = []
        for ag in df['Code_Unite'].unique():
            d = df[df['Code_Unite']==ag]
            t = len(d)
            o = (d['Statut_Final'].str.upper()=='OK').sum()
            ag_metrics.append({'Agence':ag,'Total':t,'OK':o,'KO':t-o,'Taux':round(o/t*100,2) if t else 0})
        
        df_ag = pd.DataFrame(ag_metrics)
        moy = df_ag['Taux'].mean()
        
        # Dashboard exécutif
        ws3.cell(r,1,'🎯 DASHBOARD EXÉCUTIF').font = Font(bold=True,size=13,color="FF0000")
        r += 1
        
        dash = [['Indicateur','Valeur'],
                ['🏆 Meilleure agence',f"{df_ag.loc[df_ag['Taux'].idxmax(),'Agence']} ({df_ag['Taux'].max():.1f}%)"],
                ['🔴 Pire agence',f"{df_ag.loc[df_ag['Taux'].idxmin(),'Agence']} ({df_ag['Taux'].min():.1f}%)"],
                ['📊 Taux moyen national',f'{moy:.1f}%'],
                ['⚠️ Agences en alerte (< 60%)',len(df_ag[df_ag['Taux']<60])],
                ['✅ Agences au-dessus moyenne',f"{len(df_ag[df_ag['Taux']>=moy])}/{len(df_ag)}"],
                ['📈 Total agences',len(df_ag)]]
        
        for row in dash:
            ws3.append(row)
            if dash.index(row) == 0:
                for c in ws3[r]:
                    c.fill = PatternFill(start_color="4472C4",end_color="4472C4",fill_type="solid")
                    c.font = Font(bold=True,color="FFFFFF")
            r += 1
        r += 2
        
        # Classement général
        ws3.cell(r,1,'1. 🏆 CLASSEMENT GÉNÉRAL DES AGENCES').font = Font(bold=True,size=12)
        r += 1
        
        df_ag['Écart vs Moyenne'] = (df_ag['Taux'] - moy).round(1)
        df_ag['Rang'] = df_ag['Taux'].rank(ascending=False,method='min').astype(int)
        df_ag['Statut'] = df_ag['Taux'].apply(lambda x: '🟢 Excellent' if x>=80 else '🟡 Moyen' if x>=60 else '🔴 Critique')
        df_class = df_ag.sort_values('Rang')[['Rang','Agence','Total','OK','KO','Taux','Écart vs Moyenne','Statut']]
        
        for row_idx, row in enumerate(dataframe_to_rows(df_class,index=False,header=True),r):
            for col_idx, val in enumerate(row,1):
                cell = ws3.cell(row_idx,col_idx,val)
                if row_idx == r:
                    cell.fill = PatternFill(start_color="70AD47",end_color="70AD47",fill_type="solid")
                    cell.font = Font(bold=True,color="FFFFFF")
                else:
                    if col_idx == 8:
                        if '🟢' in str(val): cell.fill = PatternFill(start_color="C6EFCE",fill_type="solid")
                        elif '🔴' in str(val): cell.fill = PatternFill(start_color="FFC7CE",fill_type="solid")
                        elif '🟡' in str(val): cell.fill = PatternFill(start_color="FFEB9C",fill_type="solid")
        r += len(df_class) + 3
        
        # Agences à risque
        risque = df_class[df_class['Taux']<60]
        if len(risque)>0:
            ws3.cell(r,1,'2. ⚠️ AGENCES À RISQUE (Taux < 60%)').font = Font(bold=True,size=12,color="C00000")
            r += 1
            risque_display = risque.copy()
            risque_display['Action recommandée'] = 'Audit urgent + Plan d\'action'
            for row_idx, row in enumerate(dataframe_to_rows(risque_display,index=False,header=True),r):
                for col_idx, val in enumerate(row,1):
                    cell = ws3.cell(row_idx,col_idx,val)
                    if row_idx == r:
                        cell.fill = PatternFill(start_color="C00000",fill_type="solid")
                        cell.font = Font(bold=True,color="FFFFFF")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE",fill_type="solid")
            r += len(risque) + 3
        
        # Top 5 performers
        top5 = df_class.head(5)
        ws3.cell(r,1,'3. 🌟 TOP 5 PERFORMERS').font = Font(bold=True,size=12,color="00B050")
        r += 1
        for row_idx, row in enumerate(dataframe_to_rows(top5,index=False,header=True),r):
            for col_idx, val in enumerate(row,1):
                cell = ws3.cell(row_idx,col_idx,val)
                if row_idx == r:
                    cell.fill = PatternFill(start_color="00B050",fill_type="solid")
                    cell.font = Font(bold=True,color="FFFFFF")
                else:
                    cell.fill = PatternFill(start_color="C6EFCE",fill_type="solid")
        r += len(top5) + 3
        
        # Volume par agence
        ws3.cell(r,1,'4. 📊 VOLUME TOTAL PAR AGENCE').font = Font(bold=True,size=12)
        r += 1
        vol = df['Code_Unite'].value_counts().reset_index()
        vol.columns = ['Agence','Nombre total']
        vol['% du total'] = round(vol['Nombre total']/total*100,2)
        for row in dataframe_to_rows(vol,index=False,header=True):
            ws3.append(row)
            r += 1
        r += 2
        
        # Croisement Agences × Types d'erreurs
        df_ko = df[df['Statut_Final'].str.upper()!='OK']
        if len(df_ko)>0:
            ws3.cell(r,1,'5. 🔀 CROISEMENT AGENCES × TYPES D\'ERREURS').font = Font(bold=True,size=12)
            r += 1
            try:
                cross = pd.crosstab(df_ko['Code_Unite'],df_ko['Statut_Final'],margins=True).reset_index()
                for row in dataframe_to_rows(cross,index=False,header=True):
                    ws3.append(row)
                    r += 1
            except: pass
    
    # ONGLET 4: Contrats OK détaillé
    if ok > 0:
        ws4 = wb.create_sheet('Contrats OK')
        df_ok = df[df['Statut_Final'].str.upper()=='OK']
        
        # Résumé
        ws4.append(['ANALYSE DES CONTRATS OK'])
        ws4.append([])
        ws4.append(['Métrique','Valeur'])
        ws4.append(['Total contrats OK',ok])
        ws4.append(['% du total',f'{round(ok/total*100,1)}%'])
        ws4.append(['Nombre de types différents',df_ok['Type (libellé)'].nunique()])
        ws4.append(['Nombre d\'agences',df_ok['Code_Unite'].nunique()])
        ws4.append([])
        
        # Par type
        ws4.append(['RÉPARTITION PAR TYPE DE CONTRAT'])
        ok_type = df_ok['Type (libellé)'].value_counts().reset_index()
        ok_type.columns = ['Type','Nombre']
        ok_type['%'] = round(ok_type['Nombre']/ok*100,1)
        for r in dataframe_to_rows(ok_type,index=False,header=True):
            ws4.append(r)
        ws4.append([])
        
        # Par agence
        ws4.append(['RÉPARTITION PAR AGENCE'])
        ok_ag = df_ok['Code_Unite'].value_counts().reset_index()
        ok_ag.columns = ['Agence','Nombre']
        ok_ag['%'] = round(ok_ag['Nombre']/ok*100,1)
        for r in dataframe_to_rows(ok_ag,index=False,header=True):
            ws4.append(r)
        
        style_ws(ws4)
    
    # ONGLET 5: Contrats KO détaillé
    if ko > 0:
        ws5 = wb.create_sheet('Contrats KO')
        df_ko = df[df['Statut_Final'].str.upper()!='OK']
        
        # Résumé
        ws5.append(['ANALYSE DES CONTRATS KO'])
        ws5.append([])
        ws5.append(['Métrique','Valeur'])
        ws5.append(['Total contrats KO',ko])
        ws5.append(['% du total',f'{round(ko/total*100,1)}%'])
        ws5.append(['Taux d\'échec',f'{round(ko/total*100,1)}%'])
        ws5.append(['Nombre de types d\'erreurs',df_ko['Statut_Final'].nunique()])
        ws5.append(['Nombre d\'agences concernées',df_ko['Code_Unite'].nunique()])
        ws5.append([])
        
        # Types d'erreurs
        ws5.append(['RÉPARTITION DES ERREURS PAR STATUT'])
        ko_stat = df_ko['Statut_Final'].value_counts().reset_index()
        ko_stat.columns = ['Type d\'erreur','Nombre']
        ko_stat['%'] = round(ko_stat['Nombre']/ko*100,1)
        for r in dataframe_to_rows(ko_stat,index=False,header=True):
            ws5.append(r)
        ws5.append([])
        
        # Par agence
        ws5.append(['REJETS PAR AGENCE'])
        ko_ag = df_ko['Code_Unite'].value_counts().reset_index()
        ko_ag.columns = ['Agence','Nombre de rejets']
        ko_ag['% des rejets'] = round(ko_ag['Nombre de rejets']/ko*100,1)
        for r in dataframe_to_rows(ko_ag,index=False,header=True):
            ws5.append(r)
        ws5.append([])
        
        # Messages d'erreur
        if 'Message_Integration' in df_ko.columns:
            msg_int = df_ko[df_ko['Message_Integration']!='']['Message_Integration'].value_counts().head(15)
            if len(msg_int)>0:
                ws5.append(['TOP 15 MESSAGES D\'ERREUR - INTÉGRATION'])
                for r in dataframe_to_rows(pd.DataFrame({'Message':msg_int.index,'Occurrences':msg_int.values}),index=False,header=True):
                    ws5.append(r)
                ws5.append([])
        
        # Par type de contrat
        ws5.append(['CONTRATS KO PAR TYPE'])
        ko_type = df_ko['Type (libellé)'].value_counts().reset_index()
        ko_type.columns = ['Type','Nombre KO']
        for r in dataframe_to_rows(ko_type,index=False,header=True):
            ws5.append(r)
        
        style_ws(ws5)
    
    # ONGLET 6: Types et Avenants
    ws6 = wb.create_sheet('Types et Avenants')
    ws6.append(['ANALYSE DES TYPES DE CONTRATS ET AVENANTS'])
    ws6.append([])
    
    # Initial vs Avenant
    ws6.append(['RÉPARTITION INITIAL VS AVENANT'])
    ia = df['Initial/Avenant'].value_counts().reset_index()
    ia.columns = ['Catégorie','Nombre']
    ia['%'] = round(ia['Nombre']/total*100,1)
    for r in dataframe_to_rows(ia,index=False,header=True):
        ws6.append(r)
    ws6.append([])
    
    # Types détaillés
    ws6.append(['DÉTAIL PAR TYPE DE CONTRAT'])
    types = df['Type (libellé)'].value_counts().reset_index()
    types.columns = ['Type','Nombre']
    types['%'] = round(types['Nombre']/total*100,1)
    for r in dataframe_to_rows(types,index=False,header=True):
        ws6.append(r)
    ws6.append([])
    
    # Croisement Type × Statut
    ws6.append(['CROISEMENT TYPE × STATUT'])
    try:
        cross_ts = pd.crosstab(df['Type (libellé)'],df['Statut_Final'],margins=True).reset_index()
        for r in dataframe_to_rows(cross_ts,index=False,header=True):
            ws6.append(r)
    except: pass
    
    style_ws(ws6)
    
    # ONGLET 7: Analyse temporelle
    if 'Date_Integration' in df.columns:
        ws7 = wb.create_sheet('Analyse temporelle')
        df_temp = df.copy()
        df_temp['Date_Integration'] = pd.to_datetime(df_temp['Date_Integration'],errors='coerce')
        df_temp = df_temp.dropna(subset=['Date_Integration'])
        
        if len(df_temp)>0:
            ws7.append(['ANALYSE TEMPORELLE'])
            ws7.append([])
            ws7.append(['Métrique','Valeur'])
            ws7.append(['Date la plus ancienne',df_temp['Date_Integration'].min().strftime('%d/%m/%Y')])
            ws7.append(['Date la plus récente',df_temp['Date_Integration'].max().strftime('%d/%m/%Y')])
            ws7.append(['Nombre de jours couverts',(df_temp['Date_Integration'].max()-df_temp['Date_Integration'].min()).days])
            ws7.append([])
            
            # Par jour
            ws7.append(['VOLUME PAR JOUR'])
            df_temp['Date'] = df_temp['Date_Integration'].dt.date
            daily = df_temp.groupby('Date').size().reset_index(name='Nombre')
            for r in dataframe_to_rows(daily,index=False,header=True):
                ws7.append(r)
            ws7.append([])
            
            # Par mois
            ws7.append(['VOLUME PAR MOIS'])
            df_temp['Mois'] = df_temp['Date_Integration'].dt.to_period('M').astype(str)
            monthly = df_temp.groupby('Mois').size().reset_index(name='Nombre')
            for r in dataframe_to_rows(monthly,index=False,header=True):
                ws7.append(r)
            
            style_ws(ws7)
    
    wb.save(output)
    output.seek(0)
    return output

# ==================== INTERFACE ====================

uploaded = st.file_uploader("📁 Fichier Excel", type=['xlsx','xls'])

if uploaded:
    try:
        df = pd.read_excel(uploaded)
        with st.expander("👁️ Aperçu", expanded=False):
            st.dataframe(df.head(10), width='stretch')
        
        df_clean = clean_data(df)
        st.success(f"✅ {len(df_clean)} lignes, {len(df_clean.columns)} colonnes")
        
        tab1,tab2,tab3,tab4,tab5,tab6 = st.tabs(["🔍 Recherche","📋 Données","🏢 Dashboard","📊 Analyses","📈 Visualisations","💾 Export"])
        
        # TAB 1: RECHERCHE
        with tab1:
            st.subheader("🔍 Recherche Hybride Intelligente")
            
            if 'hist' not in st.session_state:
                st.session_state.hist = []
            
            col1,col2 = st.columns([4,1])
            with col1:
                q = st.text_input("🔎 Recherche", placeholder="Ex: contrats ko nvm septembre")
            with col2:
                mode = st.selectbox("Mode", ["🧠 Hybride","🎯 Exact","🔤 Flou"])
            
            if q and len(q)>=2:
                sugg = get_suggestions(q, df_clean)
                if sugg:
                    with st.expander("💡 Suggestions", expanded=True):
                        cols = st.columns(min(len(sugg),5))
                        for i,s in enumerate(sugg):
                            cols[i].button(f"{s['type']}: {s['value']}", key=f"sg{i}")
            
            if st.button("🔍 RECHERCHER", type="primary") and q:
                res = df_clean.copy()
                
                if mode == "🧠 Hybride":
                    filt = parse_nl_query(q, df_clean)
                    if filt:
                        st.info(f"Filtres: {', '.join([f'{k}:{v}' for k,v in filt.items()])}")
                    
                    if filt.get('statut'):
                        res = res[res['Statut_Final'].str.upper()==filt['statut']] if filt['statut']=='OK' else res[res['Statut_Final'].str.upper()!='OK']
                    if filt.get('agence'):
                        res = res[res['Code_Unite']==filt['agence']]
                    if filt.get('mois'):
                        res['Date_Integration'] = pd.to_datetime(res['Date_Integration'],errors='coerce')
                        res = res[res['Date_Integration'].dt.month==filt['mois']]
                    
                    res['_score'] = res.apply(lambda r: calc_score(r,q,filt), axis=1)
                    res = res[res['_score']>0].sort_values('_score',ascending=False)
                
                elif mode == "🎯 Exact":
                    mask = pd.Series([False]*len(res))
                    for col in res.columns:
                        mask |= res[col].astype(str).str.contains(q,case=False,na=False)
                    res = res[mask]
                
                else:
                    if 'Contrat' in res.columns:
                        mtch = fuzzy_search(q,res,'Contrat',50)
                        if mtch:
                            res = res[res['Contrat'].isin([m[0] for m in mtch])]
                            res['_score'] = res['Contrat'].map({m[0]:m[1] for m in mtch})
                            res = res.sort_values('_score',ascending=False)
                
                if len(res)>0:
                    st.success(f"✅ {len(res)} résultat(s)")
                    if '_score' in res.columns:
                        c1,c2,c3 = st.columns(3)
                        c1.metric("Moy", f"{res['_score'].mean():.0f}%")
                        c2.metric("Max", f"{res['_score'].max():.0f}%")
                        c3.metric("Min", f"{res['_score'].min():.0f}%")
                    st.dataframe(res, width='stretch', height=400)
                    st.download_button("📥 CSV", res.to_csv(index=False).encode(), f"recherche_{datetime.now():%Y%m%d_%H%M%S}.csv")
                else:
                    st.warning(f"Aucun résultat pour '{q}'")
        
        # TAB 2: DONNÉES
        with tab2:
            st.subheader("📋 Données nettoyées")
            st.dataframe(df_clean, width='stretch', height=400)
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Lignes", len(df_clean))
            c2.metric("Colonnes", len(df_clean.columns))
            c3.metric("Doublons", df_clean.duplicated().sum())
            if 'Statut_Final' in df_clean.columns:
                c4.metric("OK", len(df_clean[df_clean['Statut_Final'].str.upper()=='OK']))
        
        # TAB 3: DASHBOARD AGENCES
        with tab3:
            st.subheader("🏢 Dashboard Agences - Vue Exécutive")
            
            if 'Code_Unite' in df_clean.columns and 'Statut_Final' in df_clean.columns:
                agm = []
                for ag in df_clean['Code_Unite'].unique():
                    d = df_clean[df_clean['Code_Unite']==ag]
                    t = len(d)
                    o = (d['Statut_Final'].str.upper()=='OK').sum()
                    agm.append({'Agence':ag,'Total':t,'OK':o,'KO':t-o,'Taux':round(o/t*100,1) if t else 0})
                
                df_ag = pd.DataFrame(agm)
                moy = df_ag['Taux'].mean()
                df_ag['Écart'] = (df_ag['Taux'] - moy).round(1)
                df_ag = df_ag.sort_values('Taux', ascending=False)
                
                # Métriques clés
                st.markdown("### 🎯 Métriques Clés")
                c1,c2,c3,c4,c5 = st.columns(5)
                c1.metric("🏆 Meilleure", df_ag.iloc[0]['Agence'], f"{df_ag.iloc[0]['Taux']}%")
                c2.metric("🔴 Pire", df_ag.iloc[-1]['Agence'], f"{df_ag.iloc[-1]['Taux']}%")
                c3.metric("📊 Moyenne", f"{moy:.1f}%")
                c4.metric("⚠️ < 60%", len(df_ag[df_ag['Taux']<60]), delta="Alerte", delta_color="inverse")
                c5.metric("✅ > Moy", f"{len(df_ag[df_ag['Taux']>=moy])}/{len(df_ag)}")
                
                # Filtres interactifs
                st.markdown("### 🔍 Filtres Interactifs")
                c1,c2,c3 = st.columns(3)
                with c1:
                    filt_ag = st.multiselect("Sélectionner agences", df_ag['Agence'].tolist(), df_ag['Agence'].tolist()[:5])
                with c2:
                    seuil = st.slider("Taux minimum (%)", 0, 100, 0)
                with c3:
                    tri = st.selectbox("Trier par", ["Taux","KO","Total","Agence"])
                
                df_f = df_ag[df_ag['Agence'].isin(filt_ag)] if filt_ag else df_ag
                df_f = df_f[df_f['Taux'] >= seuil].sort_values(tri, ascending=False)
                
                # Graphiques
                st.markdown("### 📊 Visualisations")
                c1,c2 = st.columns(2)
                
                with c1:
                    colors = ['#28a745' if x>=80 else '#ffc107' if x>=60 else '#dc3545' for x in df_f['Taux']]
                    fig = go.Figure(go.Bar(y=df_f['Agence'], x=df_f['Taux'], orientation='h',
                                          marker_color=colors, text=df_f['Taux'].apply(lambda x:f"{x:.1f}%"),
                                          textposition='outside'))
                    fig.update_layout(title="Taux de réussite par agence", xaxis_title="Taux (%)", 
                                     yaxis_title="Agence", height=400, showlegend=False)
                    st.plotly_chart(fig, use_container_width=True)
                
                with c2:
                    fig = px.scatter(df_f, x='KO', y='OK', size='Total', color='Taux',
                                    hover_name='Agence', title="Répartition OK vs KO",
                                    labels={'KO':'Rejets','OK':'Validés'}, color_continuous_scale='RdYlGn')
                    fig.update_layout(height=400)
                    st.plotly_chart(fig, use_container_width=True)
                
                # Tableau détaillé
                st.markdown("### 📋 Tableau Détaillé")
                df_f['Statut'] = df_f['Taux'].apply(lambda x: '🟢 Excellent' if x>=80 else '🟡 Moyen' if x>=60 else '🔴 Critique')
                df_f['Écart vs Moy'] = df_f['Écart'].apply(lambda x: f"{x:+.1f}%")
                st.dataframe(df_f[['Agence','Total','OK','KO','Taux','Écart vs Moy','Statut']], 
                           width='stretch', height=350, hide_index=True)
                
                # Agences à risque
                risque = df_ag[df_ag['Taux']<60]
                if len(risque)>0:
                    st.markdown("### ⚠️ Agences à Risque (< 60%)")
                    st.error(f"**{len(risque)} agence(s)** nécessite(nt) une attention immédiate")
                    c1,c2 = st.columns(2)
                    with c1:
                        st.dataframe(risque[['Agence','Taux','KO']], hide_index=True)
                    with c2:
                        st.markdown("""
                        **Actions recommandées :**
                        - 🔍 Audit approfondi des processus
                        - 📋 Plan d'action correctif urgent
                        - 👥 Formation renforcée des équipes
                        - 📊 Suivi hebdomadaire strict
                        - 💼 Support managérial
                        """)
                
                # Top Performers
                st.markdown("### 🌟 Top 5 Performers")
                top5 = df_ag.head(5)
                c1,c2 = st.columns(2)
                with c1:
                    st.dataframe(top5[['Agence','Taux','Total']], hide_index=True)
                with c2:
                    st.markdown("""
                    **Bonnes pratiques à partager :**
                    - ✅ Processus documentés et optimisés
                    - 📚 Capitalisation des connaissances
                    - 🎓 Sessions de formation inter-agences
                    - 🏆 Benchmark pour l'organisation
                    - 🤝 Mentorat des autres agences
                    """)
                
                # Évolution temporelle
                if 'Date_Integration' in df_clean.columns:
                    st.markdown("### 📈 Évolution Temporelle par Agence")
                    ag_select = st.selectbox("Sélectionner une agence", df_ag['Agence'].tolist())
                    
                    if ag_select:
                        df_temp = df_clean[df_clean['Code_Unite']==ag_select].copy()
                        df_temp['Date_Integration'] = pd.to_datetime(df_temp['Date_Integration'], errors='coerce')
                        df_temp = df_temp.dropna(subset=['Date_Integration'])
                        
                        if len(df_temp)>0:
                            df_temp['Mois'] = df_temp['Date_Integration'].dt.to_period('M').astype(str)
                            
                            monthly = []
                            for mois in df_temp['Mois'].unique():
                                dm = df_temp[df_temp['Mois']==mois]
                                t = len(dm)
                                o = (dm['Statut_Final'].str.upper()=='OK').sum()
                                monthly.append({'Mois':mois,'Total':t,'Taux':round(o/t*100,1) if t else 0})
                            
                            df_mon = pd.DataFrame(monthly).sort_values('Mois')
                            
                            fig = go.Figure()
                            fig.add_trace(go.Scatter(x=df_mon['Mois'], y=df_mon['Taux'],
                                                    mode='lines+markers', name='Taux',
                                                    line=dict(color='#4472C4',width=3),
                                                    marker=dict(size=10)))
                            fig.add_hline(y=moy, line_dash="dash", line_color="red",
                                         annotation_text=f"Moyenne: {moy:.1f}%")
                            fig.update_layout(title=f"Évolution - {ag_select}", 
                                            xaxis_title="Mois", yaxis_title="Taux (%)",
                                            height=400)
                            st.plotly_chart(fig, use_container_width=True)
                            
                            if len(df_mon)>=2:
                                tend = df_mon.iloc[-1]['Taux'] - df_mon.iloc[-2]['Taux']
                                if tend > 0:
                                    st.success(f"📈 Tendance positive : +{tend:.1f}% vs mois précédent")
                                elif tend < 0:
                                    st.error(f"📉 Tendance négative : {tend:.1f}% vs mois précédent")
                                else:
                                    st.info("→ Stable vs mois précédent")
                
                # Export dashboard
                st.markdown("### 💾 Export Dashboard")
                csv = df_f.to_csv(index=False).encode()
                st.download_button("📥 Télécharger tableau (CSV)", csv, 
                                  f"dashboard_agences_{datetime.now():%Y%m%d_%H%M%S}.csv")
            else:
                st.warning("⚠️ Colonnes 'Code_Unite' ou 'Statut_Final' manquantes")
        
        # TAB 4: ANALYSES DÉTAILLÉES
        with tab4:
            st.subheader("📊 Analyses Détaillées")
            
            # Analyse statuts
            if 'Statut_Final' in df_clean.columns:
                st.markdown("### 🎯 Analyse des Statuts")
                total = len(df_clean)
                ok_cnt = len(df_clean[df_clean['Statut_Final'].str.upper()=='OK'])
                ko_cnt = total - ok_cnt
                
                c1,c2,c3 = st.columns(3)
                c1.metric("Total contrats", total)
                c2.metric("✅ OK", ok_cnt, delta=f"{round(ok_cnt/total*100,1)}%")
                c3.metric("❌ KO", ko_cnt, delta=f"{round(ko_cnt/total*100,1)}%", delta_color="inverse")
                
                if ko_cnt > 0:
                    st.markdown("#### 🔴 Détail des Erreurs")
                    df_ko = df_clean[df_clean['Statut_Final'].str.upper()!='OK']
                    err_types = df_ko['Statut_Final'].value_counts().reset_index()
                    err_types.columns = ['Type d\'erreur','Nombre']
                    err_types['%'] = round(err_types['Nombre']/ko_cnt*100,1)
                    st.dataframe(err_types, width='stretch', hide_index=True)
            
            # Analyse Initial/Avenant
            if 'Initial/Avenant' in df_clean.columns:
                st.markdown("### 📄 Analyse Initial vs Avenants")
                ia = df_clean['Initial/Avenant'].value_counts()
                c1,c2 = st.columns(2)
                c1.metric("Contrats Initiaux", ia.get('Initial',0))
                c2.metric("Avenants", ia.get('Avenant',0))
            
            # Analyse types
            if 'Type (libellé)' in df_clean.columns:
                st.markdown("### 📋 Répartition par Type de Contrat")
                types = df_clean['Type (libellé)'].value_counts().reset_index()
                types.columns = ['Type','Nombre']
                types['%'] = round(types['Nombre']/len(df_clean)*100,1)
                st.dataframe(types, width='stretch', hide_index=True)
            
            # Croisement Agences × Erreurs
            if 'Code_Unite' in df_clean.columns and 'Statut_Final' in df_clean.columns and ko_cnt>0:
                st.markdown("### 🔀 Croisement Agences × Types d'Erreurs")
                df_ko_cross = df_clean[df_clean['Statut_Final'].str.upper()!='OK']
                try:
                    cross = pd.crosstab(df_ko_cross['Code_Unite'], df_ko_cross['Statut_Final'], 
                                       margins=True, margins_name='Total')
                    st.dataframe(cross, width='stretch')
                except:
                    st.warning("Impossible de générer le croisement")
        
        # TAB 5: VISUALISATIONS
        with tab5:
            st.subheader("📈 Visualisations Interactives")
            
            c1,c2 = st.columns(2)
            
            # Pie OK/KO
            if 'Statut_Final' in df_clean.columns:
                with c1:
                    st.markdown("#### Distribution OK vs KO")
                    ok_v = len(df_clean[df_clean['Statut_Final'].str.upper()=='OK'])
                    ko_v = len(df_clean) - ok_v
                    fig = px.pie(values=[ok_v,ko_v], names=['OK','KO'],
                                title="Répartition Statut Final", hole=0.4,
                                color_discrete_map={'OK':'#28a745','KO':'#dc3545'})
                    st.plotly_chart(fig, use_container_width=True)
            
            # Bar Types
            if 'Type (libellé)' in df_clean.columns:
                with c2:
                    st.markdown("#### Types de Contrats")
                    types_v = df_clean['Type (libellé)'].value_counts()
                    fig = px.bar(x=types_v.index, y=types_v.values,
                                title="Nombre par Type", labels={'x':'Type','y':'Nombre'},
                                color=types_v.values, color_continuous_scale='Blues')
                    fig.update_layout(showlegend=False)
                    st.plotly_chart(fig, use_container_width=True)
            
            # Graphiques Agences
            if 'Code_Unite' in df_clean.columns:
                st.markdown("#### 🏢 Analyse par Agence")
                c1,c2 = st.columns(2)
                
                with c1:
                    vol_ag = df_clean['Code_Unite'].value_counts().head(15)
                    fig = px.bar(x=vol_ag.values, y=vol_ag.index, orientation='h',
                                title="Top 15 Agences par Volume",
                                labels={'x':'Contrats','y':'Agence'},
                                color=vol_ag.values, color_continuous_scale='Viridis')
                    fig.update_layout(showlegend=False, yaxis={'categoryorder':'total ascending'})
                    st.plotly_chart(fig, use_container_width=True)
                
                with c2:
                    if 'Statut_Final' in df_clean.columns:
                        ag_succ = df_clean.groupby('Code_Unite')['Statut_Final'].apply(
                            lambda x: (x.str.upper()=='OK').sum()/len(x)*100
                        ).sort_values(ascending=False).head(15)
                        
                        fig = px.bar(x=ag_succ.values, y=ag_succ.index, orientation='h',
                                    title="Top 15 Agences - Taux de Réussite",
                                    labels={'x':'Taux (%)','y':'Agence'},
                                    color=ag_succ.values, color_continuous_scale='RdYlGn')
                        fig.update_layout(showlegend=False, yaxis={'categoryorder':'total ascending'})
                        st.plotly_chart(fig, use_container_width=True)
            
            # Timeline
            if 'Date_Integration' in df_clean.columns:
                st.markdown("#### 📅 Évolution Temporelle")
                df_time = df_clean.copy()
                df_time['Date_Integration'] = pd.to_datetime(df_time['Date_Integration'], errors='coerce')
                df_time = df_time.dropna(subset=['Date_Integration'])
                df_time['Date'] = df_time['Date_Integration'].dt.date
                timeline = df_time.groupby('Date').size().reset_index(name='Nombre')
                
                fig = px.line(timeline, x='Date', y='Nombre',
                             title="Volume de Contrats par Jour", markers=True)
                st.plotly_chart(fig, use_container_width=True)
            
            # Analyse croisée Type × Statut
            if 'Type (libellé)' in df_clean.columns and 'Statut_Final' in df_clean.columns:
                st.markdown("#### 🔀 Analyse Croisée Type × Statut")
                try:
                    cross_ts = pd.crosstab(df_clean['Type (libellé)'], df_clean['Statut_Final'])
                    fig = px.bar(cross_ts, barmode='group',
                                title="Répartition des Statuts par Type de Contrat")
                    st.plotly_chart(fig, use_container_width=True)
                except:
                    st.warning("Impossible de générer le graphique")
            
            # Heatmap Agences × Erreurs
            if 'Code_Unite' in df_clean.columns and 'Statut_Final' in df_clean.columns:
                ko_heat = df_clean[df_clean['Statut_Final'].str.upper()!='OK']
                if len(ko_heat)>0:
                    st.markdown("#### 🔥 Heatmap : Agences × Types d'Erreurs")
                    try:
                        top_ag_ko = ko_heat['Code_Unite'].value_counts().head(10).index
                        ko_heat_f = ko_heat[ko_heat['Code_Unite'].isin(top_ag_ko)]
                        heatmap = pd.crosstab(ko_heat_f['Code_Unite'], ko_heat_f['Statut_Final'])
                        
                        fig = px.imshow(heatmap, labels=dict(x="Type d'erreur",y="Agence",color="Nombre"),
                                       title="Concentration des Erreurs (Top 10 Agences)",
                                       color_continuous_scale='Reds', aspect="auto")
                        st.plotly_chart(fig, use_container_width=True)
                    except:
                        st.warning("Données insuffisantes pour la heatmap")
        
        # TAB 6: EXPORT EXCEL
        with tab6:
            st.subheader("💾 Télécharger l'Analyse Excel Complète")
            
            st.markdown("""
            ### 📑 Le fichier Excel contient 7 ONGLETS COMPLETS :
            
            1. **📄 Données nettoyées** - Toutes vos données formatées
            2. **📊 Vue d'ensemble** - Métriques clés et KPIs
            3. **🏢 Analyse par agence** - ULTRA-DÉTAILLÉE avec :
               - 🎯 Dashboard exécutif (6 indicateurs clés)
               - 🏆 Classement général avec code couleur (🟢🟡🔴)
               - ⚠️ Agences à risque (< 60%) avec actions
               - 🌟 Top 5 performers
               - 📊 Volume total par agence
               - 🔀 Croisement Agences × Types d'erreurs
            4. **✅ Contrats OK** - Analyse détaillée :
               - Résumé complet
               - Répartition par type
               - Répartition par agence
            5. **❌ Contrats KO** - Analyse approfondie :
               - Résumé des erreurs
               - Types d'erreurs détaillés
               - Rejets par agence
               - Top 15 messages d'erreur
               - KO par type de contrat
            6. **📋 Types et Avenants** :
               - Initial vs Avenant
               - Détail par type
               - Croisement Type × Statut
            7. **📅 Analyse temporelle** :
               - Volume par jour
               - Volume par mois
               - Statistiques période
            
            ### ✨ Fonctionnalités Excel :
            - 🎨 Mise en forme professionnelle automatique
            - 📊 Tableaux avec pourcentages
            - 🔍 Filtres automatiques sur toutes les feuilles
            - 📈 Codes couleur intelligents (🟢 ≥80%, 🟡 60-79%, 🔴 <60%)
            - 📏 Colonnes ajustées automatiquement
            - 🔒 En-têtes figés pour navigation facile
            - ⚠️ Alertes visuelles automatiques
            - 🏆 Classements et benchmarks
            """)
            
            excel_file = create_excel(df_clean)
            
            st.download_button(
                label="⬇️ TÉLÉCHARGER L'ANALYSE COMPLÈTE (7 ONGLETS)",
                data=excel_file,
                file_name=f"analyse_complete_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            st.success("✅ Fichier Excel ultra-détaillé prêt au téléchargement !")
            
            # Aperçu métriques
            st.markdown("### 📊 Aperçu des Métriques Clés")
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Total Contrats", len(df_clean))
            if 'Statut_Final' in df_clean.columns:
                ok_pct = round(len(df_clean[df_clean['Statut_Final'].str.upper()=='OK'])/len(df_clean)*100,1)
                c2.metric("Taux Réussite", f"{ok_pct}%")
            if 'Code_Unite' in df_clean.columns:
                c3.metric("Agences", df_clean['Code_Unite'].nunique())
            if 'Type (libellé)' in df_clean.columns:
                c4.metric("Types", df_clean['Type (libellé)'].nunique())
    
    except Exception as e:
        st.error(f"❌ Erreur : {str(e)}")
        st.exception(e)

else:
    st.info("👆 Uploadez un fichier Excel pour commencer")
    
    st.markdown("""
    ### 🚀 Excel Analyzer Pro - Fonctionnalités Complètes
    
    #### 🔍 Recherche Hybride Intelligente
    - **3 modes** : Hybride (NLP + Fuzzy), Exact, Flou
    - **Compréhension langage naturel** : "contrats ko nvm septembre"
    - **Suggestions temps réel** pendant la frappe
    - **Score de pertinence** pour trier les résultats
    - **Export CSV** des résultats
    
    #### 🏢 Dashboard Agences Interactif
    - **5 métriques clés** en temps réel
    - **Filtres dynamiques** (agences, seuil, tri)
    - **2 graphiques interactifs** avec code couleur
    - **Tableau détaillé** avec statuts visuels
    - **Alertes automatiques** agences < 60%
    - **Top 5 performers** avec bonnes pratiques
    - **Évolution temporelle** par agence
    
    #### 📊 Analyses Détaillées
    - Statistiques complètes OK/KO
    - Détail des erreurs par type
    - Répartition Initial vs Avenant
    - Types de contrats
    - Croisements multiples
    
    #### 📈 Visualisations Interactives
    - Pie charts, bar charts, line charts
    - Graphiques par agence
    - Timeline évolution
    - Heatmap erreurs
    - Analyse croisée Type × Statut
    
    #### 💾 Export Excel Ultra-Détaillé
    - **7 onglets complets** d'analyse
    - Dashboard exécutif automatique
    - Classement avec code couleur
    - Toutes les analyses incluses
    - Mise en forme professionnelle
    
    ### 📋 Formats Supportés
    - `.xlsx` (Excel 2007+)
    - `.xls` (Excel 97-2003)
    
    ### ⚡ Performance
    Optimisé pour **plusieurs dizaines de milliers de lignes**
    """)

st.markdown("---")
st.markdown("<div style='text-align:center;color:#666;'>Excel Analyzer Pro v2.0 - Solution Complète</div>", unsafe_allow_html=True)
